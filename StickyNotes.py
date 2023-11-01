import openpyxl
import os

class Reminder:
    def __init__(self, file_name):
        self.file_name = file_name
        self.load_data()

    def load_data(self):
        if os.path.exists(self.file_name):
            self.workbook = openpyxl.load_workbook(self.file_name)
            self.sheet = self.workbook.active
        else:
            self.workbook = openpyxl.Workbook()
            self.sheet = self.workbook.active
            self.sheet.append(["Number", "Date", "Time", "Notes", "Status"])
            self.workbook.save(self.file_name)

    def save_data(self):
        self.workbook.save(self.file_name)

    def list_tasks(self, status=None):
        tasks = []
        for row in self.sheet.iter_rows(values_only=True):
            if status is None or row[4] == status:
                tasks.append(row)
        return sorted(tasks, key=lambda x: (x[1], x[2]))  # Sort by Date and Time

    def mark_completed(self, task_num):
        self.sheet.cell(row=task_num + 1, column=5, value="Complete")
        self.save_data()

    def edit_task(self, task_num, date=None, time=None, notes=None, status=None):
        if date:
            self.sheet.cell(row=task_num + 1, column=2, value=date)
        if time:
            self.sheet.cell(row=task_num + 1, column=3, value=time)
        if notes:
            self.sheet.cell(row=task_num + 1, column=4, value=notes)
        if status:
            self.sheet.cell(row=task_num + 1, column=5, value=status)
        self.save_data()

    def delete_task(self, task_num):
        self.sheet.delete_rows(task_num + 1)
        self.save_data()

    def add_task(self, date, time, notes):
        num = len(self.list_tasks()) + 1
        self.sheet.append([num, date, time, notes, "Incomplete"])
        self.save_data()

def main():
    file_name = "tasks.xlsx"
    reminder = Reminder(file_name)

    while True:
        print("\nReminder Menu:")
        print("1. List Incomplete Tasks")
        print("2. List Completed Tasks")
        print("3. Mark Task as Completed")
        print("4. Edit Task")
        print("5. Add New Task")
        print("6. Delete Task")
        print("7. Exit")

        choice = input("Enter your choice: ")

        if choice == "1":
            tasks = reminder.list_tasks(status="Incomplete")
            print("Incomplete Tasks:")
            for task in tasks:
                print(f"{task[0]}. Date: {task[1]}, Time: {task[2]}, Notes: {task[3]}, Status: {task[4]}")

        elif choice == "2":
            tasks = reminder.list_tasks(status="Complete")
            print("Completed Tasks:")
            for task in tasks:
                print(f"{task[0]}. Date: {task[1]}, Time: {task[2]}, Notes: {task[3]}, Status: {task[4]}")

        elif choice == "3":
            task_num = int(input("Enter the task number to mark as completed: "))
            reminder.mark_completed(task_num - 1)
            print("Task marked as completed.")

        elif choice == "4":
            task_num = int(input("Enter the task number to edit: "))
            date = input("Enter new date (leave empty to keep current): ")
            time = input("Enter new time (leave empty to keep current): ")
            notes = input("Enter new notes (leave empty to keep current): ")
            status = input("Enter new status (leave empty to keep current): ")
            reminder.edit_task(task_num - 1, date, time, notes, status)
            print("Task edited successfully.")

        elif choice == "5":
            date = input("Enter date (MM-DD-YYYY): ")
            time = input("Enter time: ")
            notes = input("Enter notes: ")
            reminder.add_task(date, time, notes)
            print("New task added.")

        elif choice == "6":
            task_num = int(input("Enter the task number to delete: "))
            reminder.delete_task(task_num - 1)
            print("Task deleted.")

        elif choice == "7":
            print("Exiting Reminder.")
            break

        else:
            print("Invalid choice. Please enter a valid option.")

if __name__ == "__main__":
    main()

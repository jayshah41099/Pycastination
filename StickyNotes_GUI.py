import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import os

class ReminderApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Reminder App")
        self.init_ui()
        self.load_data()

    def init_ui(self):
        self.tree = ttk.Treeview(self.root, columns=("Date", "Time", "Notes", "Status"))
        self.tree.heading("#1", text="Date")
        self.tree.heading("#2", text="Time")
        self.tree.heading("#3", text="Notes")
        self.tree.heading("#4", text="Status")

        self.tree.pack()

        self.complete_button = ttk.Button(self.root, text="Change Status", command=self.change_status)
        self.complete_button.pack()

        self.edit_button = ttk.Button(self.root, text="Edit Task", command=self.edit_task)
        self.edit_button.pack()

        self.delete_button = ttk.Button(self.root, text="Delete Task", command=self.delete_task)
        self.delete_button.pack()

        self.add_button = ttk.Button(self.root, text="Add Task", command=self.add_task)
        self.add_button.pack()

    def load_data(self):
        self.tasks = []
        self.file_name = "tasksGUI.xlsx"
        if os.path.exists(self.file_name):
            workbook = openpyxl.load_workbook(self.file_name)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                self.tasks.append(list(row))
            self.update_tree()

    def save_data(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Date", "Time", "Notes", "Status"])
        for task in self.tasks:
            sheet.append(task)
        workbook.save(self.file_name)

    def update_tree(self):
        self.tree.delete(*self.tree.get_children())
        for task in self.tasks:
            self.tree.insert("", "end", values=task)

    def change_status(self):
        selected_item = self.tree.selection()
        if selected_item:
            index = int(self.tree.item(selected_item)["text"]) - 1
            status = self.tasks[index][3]
            if status == "Incomplete":
                self.tasks[index][3] = "Complete"
                self.save_data()
                self.update_tree()
            else:
                messagebox.showinfo("Info", "Task is already complete.")

    def edit_task(self):
        selected_item = self.tree.selection()
        if selected_item:
            index = int(self.tree.item(selected_item)["text"]) - 1
            edited_task = self.tasks[index]
            self.edit_window = tk.Toplevel(self.root)
            self.edit_window.title("Edit Task")
            tk.Label(self.edit_window, text="Date:").grid(row=0, column=0)
            tk.Label(self.edit_window, text="Time:").grid(row=1, column=0)
            tk.Label(self.edit_window, text="Notes:").grid(row=2, column=0)
            self.date_entry = tk.Entry(self.edit_window)
            self.date_entry.grid(row=0, column=1)
            self.date_entry.insert(0, edited_task[0])
            self.time_entry = tk.Entry(self.edit_window)
            self.time_entry.grid(row=1, column=1)
            self.time_entry.insert(0, edited_task[1])
            self.notes_entry = tk.Entry(self.edit_window)
            self.notes_entry.grid(row=2, column=1)
            self.notes_entry.insert(0, edited_task[2])
            tk.Button(self.edit_window, text="Save", command=self.save_edited_task).grid(row=3, columnspan=2)


    def save_edited_task(self):
        selected_item = self.tree.selection()
        if selected_item:
            index = int(self.tree.item(selected_item)["text"]) - 1
            self.tasks[index][0] = self.date_entry.get()
            self.tasks[index][1] = self.time_entry.get()
            self.tasks[index][2] = self.notes_entry.get()
            self.save_data()
            self.edit_window.destroy()
            self.update_tree()

    def delete_task(self):
        selected_item = self.tree.selection()
        if selected_item:
            index = int(self.tree.item(selected_item)["values"][0]) - 1
            del self.tasks[index]
            self.save_data()
            self.update_tree()

    def add_task(self):
        self.add_window = tk.Toplevel(self.root)
        self.add_window.title("Add Task")
        tk.Label(self.add_window, text="Date:").grid(row=0, column=0)
        tk.Label(self.add_window, text="Time:").grid(row=1, column=0)
        tk.Label(self.add_window, text="Notes:").grid(row=2, column=0)
        self.date_entry = tk.Entry(self.add_window)
        self.date_entry.grid(row=0, column=1)
        self.time_entry = tk.Entry(self.add_window)
        self.time_entry.grid(row=1, column=1)
        self.notes_entry = tk.Entry(self.add_window)
        self.notes_entry.grid(row=2, column=1)
        tk.Button(self.add_window, text="Add", command=self.save_new_task).grid(row=3, columnspan=2)

    def save_new_task(self):
        new_task = [
            len(self.tasks) + 1,
            self.date_entry.get(),
            self.time_entry.get(),
            self.notes_entry.get(),
            "Incomplete"
        ]
        self.tasks.append(new_task)
        self.save_data()
        self.add_window.destroy()
        self.update_tree()

def main():
    root = tk.Tk()
    app = ReminderApp(root)
    app.root.mainloop()

if __name__ == "__main__":
    main()
